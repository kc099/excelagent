import pandas as pd
import openpyxl
from openpyxl.formula.tokenizer import Tokenizer
from typing import Dict, List, Any, Tuple, Set, Optional
import logging
import networkx as nx
import re

logger = logging.getLogger(__name__)

class ExcelParser:
    """
    Core Excel parsing class that extracts data, structure and formulas from Excel files.
    """
    def __init__(self, file_path: str):
        """
        Initialize the Excel parser with the path to an Excel file.
        
        Args:
            file_path: Path to the Excel file to parse
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet_names = []
        self.named_ranges = {}
        self.tables = {}
        self.formulas = {}
        self.sheet_data = {}
        self.cell_metadata = {}
        
    def load_workbook(self) -> bool:
        """
        Load the Excel workbook using openpyxl.
        
        Returns:
            bool: True if loading was successful, False otherwise
        """
        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=False,  # We want formulas, not calculated values
                read_only=True    # Read-only for better performance with large files
            )
            self.sheet_names = self.workbook.sheetnames
            self._extract_named_ranges()
            self._extract_tables()
            return True
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return False
    
    def _extract_named_ranges(self) -> None:
        """Extract all named ranges from the workbook."""
        if self.workbook and hasattr(self.workbook, 'defined_names'):
            # In newer versions of openpyxl, defined_names is a dictionary-like object
            # that you can iterate over directly
            for name, defn in self.workbook.defined_names.items():
                destinations = defn.destinations
                for sheet, coord in destinations:
                    self.named_ranges[name] = f"{sheet}!{coord}"
    
    def _extract_tables(self) -> None:
        """Extract all tables from the workbook."""
        if self.workbook:
            for sheet_name in self.sheet_names:
                sheet = self.workbook[sheet_name]
                if hasattr(sheet, '_tables'):
                    for table_name, table_obj in sheet._tables.items():
                        self.tables[table_name] = {
                            'sheet': sheet_name,
                            'ref': table_obj.ref,
                            'display_name': table_obj.displayName,
                            'header_row_count': table_obj.headerRowCount
                        }
    
    def parse_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        Parse all sheets in the workbook into pandas DataFrames.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary mapping sheet names to DataFrames
        """
        if not self.workbook:
            if not self.load_workbook():
                return {}
        
        for sheet_name in self.sheet_names:
            self.parse_sheet(sheet_name)
        
        return self.sheet_data
    
    def parse_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """
        Parse a single sheet into a pandas DataFrame.
        
        Args:
            sheet_name: Name of the sheet to parse
            
        Returns:
            Optional[pd.DataFrame]: DataFrame containing the sheet data, or None if the sheet doesn't exist
        """
        if not self.workbook:
            if not self.load_workbook():
                return None
        
        if sheet_name not in self.sheet_names:
            logger.warning(f"Sheet {sheet_name} not found in workbook")
            return None
        
        sheet = self.workbook[sheet_name]
        data = []
        max_col = 0
        
        # Track formulas
        sheet_formulas = {}
        
        # Process each row
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            row_data = []
            for col_idx, cell in enumerate(row, 1):
                value = cell.value
                
                # Check if cell has coordinate attribute (EmptyCell objects don't have it)
                if hasattr(cell, 'coordinate'):
                    cell_addr = f"{sheet_name}!{cell.coordinate}"
                else:
                    # Create coordinate for empty cell using column letter and row number
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx)
                    cell_addr = f"{sheet_name}!{col_letter}{row_idx}"
                
                # Store cell metadata
                self.cell_metadata[cell_addr] = {
                    'sheet': sheet_name,
                    'row': row_idx,
                    'column': col_idx,
                    'data_type': cell.data_type,
                    'style_id': cell.style_id if hasattr(cell, 'style_id') else None
                }
                
                # Extract formula if present
                if cell.data_type == 'f':
                    formula = cell.value
                    if formula and formula.startswith('='):
                        sheet_formulas[cell_addr] = formula
                        self._parse_formula(sheet_name, cell_addr, formula)
                
                row_data.append(value)
            
            if row_data:
                max_col = max(max_col, len(row_data))
                data.append(row_data)
        
        # Update formulas dictionary
        if sheet_formulas:
            self.formulas[sheet_name] = sheet_formulas
        
        # Create DataFrame with the appropriate number of columns
        if data:
            # Ensure all rows have the same number of columns
            for i, row in enumerate(data):
                if len(row) < max_col:
                    data[i] = row + [None] * (max_col - len(row))
            
            df = pd.DataFrame(data)
            
            # Try to use the first row as headers if it seems to contain headers
            if len(df) > 0:
                # Check if first row seems like headers (e.g., string values)
                if all(isinstance(x, str) for x in df.iloc[0] if x is not None):
                    df.columns = df.iloc[0]
                    df = df.iloc[1:].reset_index(drop=True)
            
            self.sheet_data[sheet_name] = df
            return df
        
        return pd.DataFrame()
    
    def _parse_formula(self, sheet_name: str, cell_addr: str, formula: str) -> None:
        """
        Parse a formula to extract dependencies and structure.
        
        Args:
            sheet_name: Name of the sheet containing the formula
            cell_addr: Cell address (e.g., 'Sheet1!A1')
            formula: Formula string
        """
        try:
            # Remove the leading equals sign
            formula_text = formula[1:] if formula.startswith('=') else formula
            
            # Use openpyxl's tokenizer to parse the formula
            tokens = Tokenizer(formula_text).items
            
            # Extract cell references from the formula
            cell_refs = []
            for token in tokens:
                if token.type == 'CELL':
                    # If it's a reference to another sheet
                    if '!' in token.value:
                        cell_refs.append(token.value)
                    else:
                        # If it's a reference in the same sheet
                        cell_refs.append(f"{sheet_name}!{token.value}")
                elif token.type == 'RANGE':
                    # Handle range references
                    if '!' in token.value:
                        cell_refs.append(token.value)
                    else:
                        cell_refs.append(f"{sheet_name}!{token.value}")
            
            # Store formula metadata
            if sheet_name not in self.formulas:
                self.formulas[sheet_name] = {}
            
            self.formulas[sheet_name][cell_addr] = {
                'formula': formula,
                'dependencies': cell_refs,
                'tokens': [{'type': t.type, 'value': t.value} for t in tokens]
            }
            
        except Exception as e:
            logger.warning(f"Error parsing formula in {cell_addr}: {e}")
    
    def get_formula_dependencies(self) -> Dict[str, List[str]]:
        """
        Get all formula dependencies as a dictionary.
        
        Returns:
            Dict[str, List[str]]: Dictionary mapping cell addresses to their dependencies
        """
        dependencies = {}
        
        for sheet_name, sheet_formulas in self.formulas.items():
            for cell_addr, formula_data in sheet_formulas.items():
                if isinstance(formula_data, dict) and 'dependencies' in formula_data:
                    dependencies[cell_addr] = formula_data['dependencies']
                elif isinstance(formula_data, str):
                    # For backward compatibility
                    dependencies[cell_addr] = []
        
        return dependencies
    
    def categorize_formulas(self) -> Dict[str, str]:
        """
        Categorize formulas by their function type.
        
        Returns:
            Dict[str, str]: Dictionary mapping cell addresses to formula categories
        """
        categories = {}
        
        # Function patterns for categorization
        patterns = {
            'math': re.compile(r'\b(SUM|AVERAGE|MIN|MAX|COUNT|PRODUCT|ROUND|INT|ABS)\(', re.IGNORECASE),
            'logical': re.compile(r'\b(IF|AND|OR|NOT|TRUE|FALSE|IFERROR|IFNA)\(', re.IGNORECASE),
            'text': re.compile(r'\b(CONCATENATE|LEFT|RIGHT|MID|LEN|LOWER|UPPER|PROPER|TRIM|TEXT)\(', re.IGNORECASE),
            'lookup': re.compile(r'\b(VLOOKUP|HLOOKUP|LOOKUP|MATCH|INDEX|INDIRECT|OFFSET)\(', re.IGNORECASE),
            'date': re.compile(r'\b(TODAY|NOW|DATE|DATEDIF|DATEVALUE|DAY|MONTH|YEAR|WEEKDAY)\(', re.IGNORECASE),
            'financial': re.compile(r'\b(NPV|IRR|PMT|FV|PV|RATE|IPMT|PPMT)\(', re.IGNORECASE),
            'statistical': re.compile(r'\b(STDEV|VAR|CORREL|FORECAST|TREND|PERCENTILE|QUARTILE|MEDIAN)\(', re.IGNORECASE)
        }
        
        for sheet_name, sheet_formulas in self.formulas.items():
            for cell_addr, formula_data in sheet_formulas.items():
                formula = ""
                if isinstance(formula_data, dict) and 'formula' in formula_data:
                    formula = formula_data['formula']
                elif isinstance(formula_data, str):
                    formula = formula_data
                
                # Determine the category
                category = 'other'
                for cat_name, pattern in patterns.items():
                    if pattern.search(formula):
                        category = cat_name
                        break
                
                categories[cell_addr] = category
        
        return categories
    
    def get_sheet_metadata(self) -> Dict[str, Dict[str, Any]]:
        """
        Get metadata for all sheets.
        
        Returns:
            Dict[str, Dict[str, Any]]: Dictionary with sheet metadata
        """
        metadata = {}
        
        if not self.workbook:
            if not self.load_workbook():
                return metadata
        
        for sheet_name in self.sheet_names:
            sheet = self.workbook[sheet_name]
            
            metadata[sheet_name] = {
                'title': sheet.title,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'sheet_state': sheet.sheet_state,
                'sheet_view': {
                    'view': sheet.sheet_view.view,
                    'zoom_scale': sheet.sheet_view.zoomScale
                } if hasattr(sheet, 'sheet_view') else {}
            }
        
        return metadata
    
    def get_cell_styles(self, sheet_name: str = None) -> Dict[str, Dict[str, Any]]:
        """
        Get styles for cells in the specified sheet or all sheets.
        
        Args:
            sheet_name: Optional name of sheet to get styles for
            
        Returns:
            Dict[str, Dict[str, Any]]: Dictionary mapping cell addresses to style information
        """
        styles = {}
        
        if not self.workbook:
            if not self.load_workbook():
                return styles
        
        sheets_to_process = [sheet_name] if sheet_name else self.sheet_names
        
        for sheet_name in sheets_to_process:
            if sheet_name not in self.sheet_names:
                continue
                
            sheet = self.workbook[sheet_name]
            
            # Process only non-empty cells to save memory
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value is not None:
                        cell_addr = f"{sheet_name}!{cell.coordinate}"
                        
                        # Extract style information
                        font = cell.font if hasattr(cell, 'font') else None
                        fill = cell.fill if hasattr(cell, 'fill') else None
                        border = cell.border if hasattr(cell, 'border') else None
                        alignment = cell.alignment if hasattr(cell, 'alignment') else None
                        number_format = cell.number_format if hasattr(cell, 'number_format') else None
                        
                        styles[cell_addr] = {
                            'font': {
                                'name': font.name if font and hasattr(font, 'name') else None,
                                'size': font.size if font and hasattr(font, 'size') else None,
                                'bold': font.bold if font and hasattr(font, 'bold') else False,
                                'italic': font.italic if font and hasattr(font, 'italic') else False,
                                'color': font.color.rgb if font and hasattr(font, 'color') and hasattr(font.color, 'rgb') else None
                            } if font else {},
                            'fill': {
                                'fill_type': fill.fill_type if fill and hasattr(fill, 'fill_type') else None,
                                'start_color': fill.start_color.rgb if fill and hasattr(fill, 'start_color') and hasattr(fill.start_color, 'rgb') else None,
                                'end_color': fill.end_color.rgb if fill and hasattr(fill, 'end_color') and hasattr(fill.end_color, 'rgb') else None
                            } if fill else {},
                            'number_format': number_format,
                            'alignment': {
                                'horizontal': alignment.horizontal if alignment and hasattr(alignment, 'horizontal') else None,
                                'vertical': alignment.vertical if alignment and hasattr(alignment, 'vertical') else None
                            } if alignment else {}
                        }
        
        return styles